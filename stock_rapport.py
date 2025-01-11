
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook


data={
    "year":[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16], 

}

# Last inn Excel-filen

for year in range(0,4):

    filnavn = f"Telenor_rapport/Q4-202{str(year)}.xlsx"
    workbook = load_workbook(filnavn)

    print("\n")

    print(f"Henter data fra rapport 202{str(year)}")
    print("____________________________")
    
    # Loop gjennom alle ark i filen
    for sheet_name in workbook.sheetnames:
        ark = workbook[sheet_name]
       
        if sheet_name!="Analytical information" and sheet_name!="Analytical_information" :
            
            # Eksempel: Iterer gjennom rader i arket
            for rad in ark.iter_rows(values_only=True):
                ekstra_forsyvning=1
                if type(rad[0])==str:
                    if "Average revenue per subscription per month (ARPU) in the quarter" in rad[0]:
                        verdi=None
                        ekstra_forsyvning=1
                        while True:
                            if rad[-4:][-1]!=None:
                                verdi=rad[-4:]
                                break


                            if rad[-4-ekstra_forsyvning:-ekstra_forsyvning][-1]==None:
                                ekstra_forsyvning+=1
                            else:
                                verdi=rad[-4-ekstra_forsyvning:-ekstra_forsyvning]
                                
                                break
                        #print(f"Data fra ark: {sheet_name}")
                        try:
                            data[sheet_name]+=list(verdi)
                        except KeyError:
                            data[sheet_name]=list(verdi)
                        
                        

max_len=0
land=[]


for i in data:
    land.append(i)
    if len(data[i]) > max_len:
        max_len=len(data[i]) 

for i in data:
    if len(data[i])<max_len:
        antall_lagttil_nuller=max_len-len(data[i])
        null_liste=[0]*antall_lagttil_nuller
        data[i]=null_liste+data[i]




df = pd.DataFrame(data)

# Plot flere linjer
df.plot(x='year', y=land[1:], kind='line', title='Average revenue per subscription per month (ARPU) in the quarter')

# Vis diagrammet
plt.show()       #ARPU




















